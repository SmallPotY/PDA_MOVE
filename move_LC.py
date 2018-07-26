import requests
import openpyxl
import json

class Exceldb:

    def __init__(self):
        self.wb = openpyxl.load_workbook('move.xlsx')  # 打开excel文件
        self.sheet = self.wb.active  # 获取激活的工作表
        self.max_column = self.sheet.max_column  # 最大列数
        self.max_row = self.sheet.max_row  # 最大行数
        self.row = 2

    def next(self):
        self.item = [i.value for i in self.sheet[self.row]]
        self.row += 1
        return self.item


class Move_Lc:

    def __init__(self):
        print('读取配置文件....')
        with open('config.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
            print('移位账号：',data['user'])
            print('服务器地址:',data['app_ip'])
            print('仓库代码',data['app_db'])

        self.user = data['user']
        self.password = data['password']
        self.app_ip = data['app_ip']
        self.app_db = data['app_db']


    def login(self):
        login_headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Content-Length': '43',
            'Content-Type': 'application/x-www-form-urlencoded',
            'Host': self.app_ip,
            'Pragma': 'no-cache',
            'Referer': 'http://{}/SCV_RF/login/LoginHandling.aspx'.format(self.app_ip),
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:61.0) Gecko/20100101 Firefox/61.0'
        }

        login_data = {
            'hdConnStr': '',
            'hdSessID': '',
            'tbPassword': self.password,
            'tbUser': self.user
        }

        login_url = 'http://{}/SCV_RF/login/LoginHandling.aspx'.format(self.app_ip)


        response = requests.post(url=login_url, data=login_data, headers=login_headers)

        cookies = response.cookies

        login_headers = {
            'Host': self.app_ip,
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:61.0) Gecko/20100101 Firefox/61.0',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
            'Accept-Encoding': 'gzip, deflate',
            'Referer': 'http://{}/SCV_RF/login/LoginHandling.aspx'.format(self.app_ip),
            'Content-Type': 'application/x-www-form-urlencoded',
            'Content-Length': '418',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Pragma': 'no-cache',
            'Cache-Control': 'no-cache'
        }

        login_data = "tbUser={user}&tbPassword={password}&tbEnv=Data+Source%3D%28local%29%3BInitial+Catalog%3D{db}%3BUser+ID%3Dvip%3BPassword%3Dvip2015%3BPooling%3DTrue%3BMin+Pool+Size%3D20%3BMax+Pool+Size%3D2000%3BConnection+Timeout%3D10%3B&hdSessID=whatever&hdConnStr=Data+Source%3D%28local%29%3BInitial+Catalog%3D{db}%3BUser+ID%3Dvip%3BPassword%3Dvip2015%3BPooling%3DTrue%3BMin+Pool+Size%3D20%3BMax+Pool+Size%3D2000%3BConnection+Timeout%3D10%3B".format(user=self.user,password=self.password,db=self.app_db)
        login_url = 'http://{}/SCV_RF/login/LoginHandling.aspx'.format(self.app_ip)

        response = requests.post(url=login_url, data=login_data, headers=login_headers, cookies=cookies)

        # cookies = response.cookies
        return cookies


    def move(self,cookies,tbFromLoc,tbToLoc):
        move_headers = {
            'Host': self.app_ip,
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:61.0) Gecko/20100101 Firefox/61.0',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
            'Accept-Encoding': 'gzip, deflate',
            'Referer': 'http://58.252.75.38:8091/SCV_RF/whinvtransfer/ToLocEntry.aspx',
            'Content-Type': 'application/x-www-form-urlencoded',
            'Content-Length': '27',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            # 'Cookie': "ASP.NET_SessionId=lnbcao55r01omt3vk2vwbunl",
            'Pragma': 'no-cache',
            'Cache-Control': 'no-cache'
        }

        move_url = 'http://{}/SCV_RF/whinvtransfer/ToLocEntryHandling.aspx'.format(self.app_ip)

        move = 'tbFromLoc={tbFromLoc}&tbToLoc={tbToLoc}'.format(tbFromLoc=tbFromLoc,tbToLoc=tbToLoc)
        # print(move)

        re = requests.post(url=move_url, data=move, headers=move_headers, cookies=cookies)

        # print(re.text)

        if '请输入到货位' in re.text:
            if "错误：从库位存在移入数量或分配数量不为零的库存" in re.text:
                print('存在已分配库存，移位失败')
            else:
                print('移位成功')
        else:
            print('cookies过期，移位失败')


if __name__ == '__main__':

    obj = Move_Lc()

    cookies = obj.login()

    print('PDA模拟登陆成功')
    print('*'*15)
    # obj.move(cookies=cookies,tbFromLoc='GA01-04',tbToLoc='GA01-05')

    db = Exceldb()

    for i in range(db.max_row - 1):
        item = db.next()
        print('准备移位，从{}移动至{}'.format(item[0],item[1]))
        obj.move(cookies=cookies, tbFromLoc=item[0], tbToLoc=item[1])
        print('-'*15)


    print('\n\n\n')
    input('任务完成，按任意键关闭')